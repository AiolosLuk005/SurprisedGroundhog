from __future__ import annotations

import argparse
import csv
import os
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

from core.normalize_base import NormalizeResult, register
from core.settings import SETTINGS

TIMEOUT = SETTINGS.get("normalize", {}).get("timeouts", {}).get("excel", 3600)


def process_excel(in_path: str, out_dir: Path) -> None:
    wb = load_workbook(in_path, data_only=True)
    out_dir.mkdir(parents=True, exist_ok=True)
    md_lines = []
    for idx, sheet in enumerate(wb.worksheets, start=1):
        csv_path = out_dir / f"table_{idx}.csv"
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(["" if v is None else v for v in row])
        md_lines.append(f"sheet {idx}: {sheet.title}\n")
    (out_dir / "document.md").write_text("\n".join(md_lines), encoding="utf-8")


def cli() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()
    process_excel(args.input, Path(args.out))
    return 0


class ExcelNormalizer:
    name = "excel"
    priority = 90
    exts = ["xlsx", "xlsm", "xls"]

    def can_handle(self, path: str) -> bool:
        return any(path.lower().endswith("." + e) for e in self.exts)

    def normalize(self, path: str, out_root: str) -> NormalizeResult:
        cmd = [sys.executable, str(Path(__file__).resolve()), "--input", path, "--out", out_root]
        env = os.environ.copy()
        env.setdefault("PYTHONPATH", str(Path(__file__).resolve().parents[2]))
        try:
            proc = subprocess.run(cmd, capture_output=True, text=True, timeout=TIMEOUT, env=env)
        except Exception as e:
            return NormalizeResult(False, message=str(e))
        if proc.returncode != 0:
            msg = (proc.stderr or proc.stdout or "")[-800:]
            return NormalizeResult(False, message=msg)
        md_paths = [str(p) for p in Path(out_root).glob("*.md")]
        csv_paths = [str(p) for p in Path(out_root).glob("*.csv")]
        sidecar = str(Path(out_root) / "sidecar.json") if (Path(out_root) / "sidecar.json").exists() else None
        return NormalizeResult(True, md_paths=md_paths, csv_paths=csv_paths, sidecar=sidecar)


register(ExcelNormalizer())

if __name__ == "__main__":
    raise SystemExit(cli())
